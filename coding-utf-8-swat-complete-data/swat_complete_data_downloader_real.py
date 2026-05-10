# -*- coding: utf-8 -*-
"""
SWAT COMPLETE DATA DOWNLOADER - REAL OBSERVED DATA

Downloads and prepares:
  - IMD rainfall and temperature with watershed area weighting
  - NASA POWER solar radiation, wind speed, and relative humidity

Required packages:
    pip install imdlib xarray netcdf4 pandas numpy geopandas shapely
    pip install requests openpyxl xlsxwriter scipy rasterio rioxarray

Notes:
    - Edit the USER CONFIGURATION section before running.
"""

from __future__ import annotations

import os
import sys
import time
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import geopandas as gpd
import xarray as xr
import requests
from shapely.geometry import Point, box

# --------------------------------------------------------------------------- #
# Optional heavy dependencies
# --------------------------------------------------------------------------- #
try:
    import rasterio
    from rasterio.mask import mask

    HAS_RASTERIO = True
except ImportError:
    HAS_RASTERIO = False
    rasterio = None
    mask = None

try:
    import imdlib as imd

    HAS_IMDLIB = True
except ImportError:
    HAS_IMDLIB = False
    imd = None

# ============================================================================ #
# USER CONFIGURATION - EDIT THESE VALUES
# ============================================================================ #

WATERSHED_SHP = r"C:\SWAT_Project\GIS\Watershed_Boundary.shp"
DEM_PATH = r"C:\SWAT_Project\GIS\DEM.tif"
OUTPUT_DIR = r"C:\SWAT_Project\Thoubal_SWAT_Real"
IMD_RAW_FOLDER = r"C:\SWAT_Project\IMD_Raw_Data"

START_YEAR = 2010
END_YEAR = 2010

NUM_ELEVATION_BANDS = 5
STATIONS_PER_BAND = 1

# Lapse rates for Thoubal River Catchment
TEMP_LAPSE_RATE = -6.5      # deg C / 1000 m
PRECIP_LAPSE_RATE = 16.0    # percent / 100 m above mean or base elevation
MANUAL_ELEVATION = 780.0    # fallback mean elevation if DEM fails

# IMD grid resolutions
IMD_RAINFALL_RES = 0.25
IMD_TEMP_RES = 1.0


# ============================================================================ #
# SMALL HELPERS
# ============================================================================ #

def line(title: str = "", char: str = "=") -> None:
    print("\n" + char * 80)
    if title:
        print(title)
        print(char * 80)


def ok(message: str) -> None:
    print(f"    [OK] {message}")


def warn(message: str) -> None:
    print(f"    [WARN] {message}")


def fail(message: str) -> None:
    print(f"    [FAIL] {message}")


def geometry_union(gdf: gpd.GeoDataFrame):
    """Return a single union geometry across GeoPandas versions."""
    try:
        return gdf.geometry.union_all()
    except AttributeError:
        return gdf.unary_union


def require_crs(gdf: gpd.GeoDataFrame, label: str) -> None:
    """Fail fast if CRS is missing; this avoids silently wrong reprojections."""
    if getattr(gdf, "crs", None) is None:
        raise ValueError(
            f"{label} has no CRS defined. Define/assign the correct CRS in GIS "
            f"(QGIS/ArcGIS) and re-save the shapefile before running."
        )


def ensure_datetime_column(df: pd.DataFrame, column: str = "time") -> pd.DataFrame:
    df = df.copy()
    df[column] = pd.to_datetime(df[column])
    return df


def safe_filename_stem(value: str, fallback: str) -> str:
    """Return a Windows-safe file-name stem without changing SWAT file formats."""
    cleaned = "".join(ch if ch.isalnum() or ch in (" ", "_", "-") else "_" for ch in value.strip())
    cleaned = "_".join(cleaned.split())
    return cleaned or fallback


# ============================================================================ #
# IMD GRID CELL
# ============================================================================ #

@dataclass
class IMDGridCell:
    """One IMD grid cell with geometry, data and area-weight."""

    lat: float
    lon: float
    resolution: float
    data: Optional[pd.DataFrame] = None
    weight: float = 0.0
    polygon: object = field(init=False)

    def __post_init__(self) -> None:
        self.polygon = box(
            self.lon,
            self.lat,
            self.lon + self.resolution,
            self.lat + self.resolution,
        )


# ============================================================================ #
# DEM PROCESSOR
# ============================================================================ #

class DEMProcessor:
    """DEM loading, masking and elevation-band creation."""

    def __init__(self, dem_path: str):
        self.dem_path = dem_path
        self.dem_crs = None

    def load_dem(self) -> bool:
        print("\n  Loading DEM...")
        if not HAS_RASTERIO:
            warn("rasterio is not installed. DEM processing will use defaults.")
            return False
        if not os.path.exists(self.dem_path):
            warn(f"DEM file not found: {self.dem_path}")
            return False
        try:
            with rasterio.open(self.dem_path) as src:
                self.dem_crs = src.crs
            ok(f"DEM loaded. CRS = {self.dem_crs}")
            return True
        except Exception as exc:
            fail(f"DEM error: {exc}")
            return False

    def _valid_elevation_values(self, elev: np.ndarray, nodata) -> np.ndarray:
        values = elev.astype(float).ravel()
        values = values[np.isfinite(values)]
        if nodata is not None and np.isfinite(nodata):
            values = values[values != float(nodata)]
        # Guard against common fill values that appear even without nodata metadata.
        values = values[(values > -500.0) & (values < 9000.0)]
        return values

    def extract_watershed_stats(self, watershed_gdf: gpd.GeoDataFrame) -> Optional[dict]:
        print("\n  Extracting elevation statistics...")
        if not HAS_RASTERIO or self.dem_crs is None:
            return None
        try:
            require_crs(watershed_gdf, "Watershed layer")
            ws = watershed_gdf.to_crs(self.dem_crs) if watershed_gdf.crs != self.dem_crs else watershed_gdf
            with rasterio.open(self.dem_path) as src:
                out_image, _ = mask(src, ws.geometry, crop=True)
                values = self._valid_elevation_values(out_image[0], src.nodata)
            if values.size == 0:
                fail("No valid elevation pixels inside watershed.")
                return None
            stats = {
                "min": float(values.min()),
                "max": float(values.max()),
                "mean": float(values.mean()),
                "range": float(values.max() - values.min()),
            }
            ok(
                f"Min: {stats['min']:.0f} m  Max: {stats['max']:.0f} m  "
                f"Mean: {stats['mean']:.0f} m"
            )
            return stats
        except Exception as exc:
            fail(f"Elevation statistics error: {exc}")
            return None

    def extract_point_elevation(self, lat: float, lon: float) -> Optional[float]:
        if not HAS_RASTERIO or self.dem_crs is None:
            return None
        try:
            pt_gdf = gpd.GeoDataFrame(geometry=[Point(lon, lat)], crs="EPSG:4326")
            pt_proj = pt_gdf.to_crs(self.dem_crs)
            with rasterio.open(self.dem_path) as src:
                xy = [(pt_proj.geometry.x.iloc[0], pt_proj.geometry.y.iloc[0])]
                values = list(src.sample(xy))
                if values:
                    value = float(values[0][0])
                    if src.nodata is not None and value == float(src.nodata):
                        return None
                    if np.isfinite(value) and -500.0 < value < 9000.0:
                        return value
        except Exception:
            pass
        return None

    def create_elevation_bands(
        self,
        watershed_gdf: gpd.GeoDataFrame,
        num_bands: int = 5,
    ) -> pd.DataFrame:
        print(f"\n  Creating {num_bands} elevation bands...")
        if not HAS_RASTERIO or self.dem_crs is None:
            return self._default_bands(num_bands)
        try:
            require_crs(watershed_gdf, "Watershed layer")
            ws = watershed_gdf.to_crs(self.dem_crs) if watershed_gdf.crs != self.dem_crs else watershed_gdf
            with rasterio.open(self.dem_path) as src:
                out_image, _ = mask(src, ws.geometry, crop=True)
                values = self._valid_elevation_values(out_image[0], src.nodata)
            if values.size == 0:
                return self._default_bands(num_bands)

            edges = np.percentile(values, np.linspace(0, 100, num_bands + 1))
            mean_elev = float(values.mean())
            records = []
            for i in range(num_bands):
                lo = float(edges[i])
                hi = float(edges[i + 1])
                centre = (lo + hi) / 2.0
                if i == num_bands - 1:
                    in_band = (values >= lo) & (values <= hi)
                else:
                    in_band = (values >= lo) & (values < hi)
                n_px = int(np.sum(in_band))
                records.append(
                    {
                        "Band": i + 1,
                        "Elev_Min_m": lo,
                        "Elev_Max_m": hi,
                        "Elev_Center_m": centre,
                        "Area_Fraction": float(n_px / values.size),
                        "Temp_Adjustment_C": float(TEMP_LAPSE_RATE * (centre - mean_elev) / 1000.0),
                        "Precip_Multiplier": float(
                            1.0 + (PRECIP_LAPSE_RATE / 100.0) * max(0.0, centre - edges[0]) / 100.0
                        ),
                    }
                )
                print(f"    Band {i + 1}: {lo:.0f}-{hi:.0f} m ({records[-1]['Area_Fraction'] * 100:.1f}%)")
            return pd.DataFrame(records)
        except Exception as exc:
            fail(f"Elevation band error: {exc}")
            return self._default_bands(num_bands)

    def _default_bands(self, num_bands: int) -> pd.DataFrame:
        warn("Using fallback elevation bands.")
        base = MANUAL_ELEVATION - 200.0 * (num_bands // 2)
        return pd.DataFrame(
            [
                {
                    "Band": i + 1,
                    "Elev_Min_m": base + i * 200.0,
                    "Elev_Max_m": base + (i + 1) * 200.0,
                    "Elev_Center_m": base + i * 200.0 + 100.0,
                    "Area_Fraction": 1.0 / num_bands,
                    "Temp_Adjustment_C": 0.0,
                    "Precip_Multiplier": 1.0,
                }
                for i in range(num_bands)
            ]
        )


# ============================================================================ #
# AREA-WEIGHTED IMD DOWNLOADER
# ============================================================================ #

class AreaWeightedIMDDownloader:
    """Download IMD gridded rainfall and temperature with area-weighting."""

    def __init__(self, raw_folder: str, start_year: int, end_year: int):
        self.raw_folder = raw_folder
        self.start_year = start_year
        self.end_year = end_year
        os.makedirs(raw_folder, exist_ok=True)
        self.rainfall_cells: list[IMDGridCell] = []
        self.temp_cells: list[IMDGridCell] = []
        self._cache: dict[tuple[str, int], xr.Dataset] = {}

    def _load_imd_xarray(self, variable: str, year: int) -> xr.Dataset:
        key = (variable, year)
        if key not in self._cache:
            if not HAS_IMDLIB:
                raise RuntimeError("imdlib is not installed. Run: pip install imdlib")
            data = imd.get_data(variable, year, year, fn_format="yearwise", file_dir=self.raw_folder)
            self._cache[key] = data.get_xarray()
        return self._cache[key]

    def _find_cells(self, watershed_gdf: gpd.GeoDataFrame, resolution: float) -> list[IMDGridCell]:
        require_crs(watershed_gdf, "Watershed layer")
        ws4326 = watershed_gdf.to_crs("EPSG:4326")
        bounds = ws4326.total_bounds
        ws_poly4326 = geometry_union(ws4326)

        ws_equal_area = ws4326.to_crs("EPSG:6933")
        ws_poly_equal_area = geometry_union(ws_equal_area)
        ws_area = float(ws_equal_area.area.sum())

        print(f"    Watershed bounds (EPSG:4326):")
        print(f"      lon_min={bounds[0]:.5f}, lat_min={bounds[1]:.5f}, lon_max={bounds[2]:.5f}, lat_max={bounds[3]:.5f}")
        print(f"    Target IMD grid resolution: {resolution} deg")

        lat_min = np.floor(bounds[1] / resolution) * resolution
        lat_max = np.ceil(bounds[3] / resolution) * resolution
        lon_min = np.floor(bounds[0] / resolution) * resolution
        lon_max = np.ceil(bounds[2] / resolution) * resolution

        n_lat = max(0, int(np.round((lat_max - lat_min) / resolution)))
        n_lon = max(0, int(np.round((lon_max - lon_min) / resolution)))
        print(f"    Grid scan window:")
        print(f"      lat: {lat_min:.5f} to {lat_max:.5f} step {resolution}  (rows ~ {n_lat})")
        print(f"      lon: {lon_min:.5f} to {lon_max:.5f} step {resolution}  (cols ~ {n_lon})")

        cells: list[IMDGridCell] = []
        lat = lat_min
        while lat < lat_max:
            lon = lon_min
            while lon < lon_max:
                cell = IMDGridCell(float(lat), float(lon), float(resolution))
                if ws_poly4326.intersects(cell.polygon):
                    cell_gdf = gpd.GeoDataFrame(geometry=[cell.polygon], crs="EPSG:4326").to_crs("EPSG:6933")
                    inter = ws_poly_equal_area.intersection(cell_gdf.geometry.iloc[0])
                    if not inter.is_empty and ws_area > 0:
                        cell.weight = float(inter.area / ws_area)
                        if cell.weight > 0:
                            cells.append(cell)
                            print(f"    Cell ({lat:.2f}, {lon:.2f}): {cell.weight * 100:.1f}% overlap")
                lon += resolution
            lat += resolution

        total = sum(c.weight for c in cells)
        if total > 0:
            for c in cells:
                c.weight /= total
        if cells:
            weights = np.array([c.weight for c in cells], dtype=float)
            print(f"    Weight normalisation: sum={weights.sum():.6f}, min={weights.min():.6f}, max={weights.max():.6f}")
        else:
            print("    No overlapping IMD cells found for this resolution.")
        return cells

    def identify_overlapping_cells(self, watershed_gdf: gpd.GeoDataFrame) -> None:
        line("[AREA-WEIGHTED] IDENTIFYING OVERLAPPING IMD GRID CELLS")
        print(f"\n  IMD RAINFALL GRID ({IMD_RAINFALL_RES} deg):")
        self.rainfall_cells = self._find_cells(watershed_gdf, IMD_RAINFALL_RES)
        print(f"\n    Total rainfall cells: {len(self.rainfall_cells)}")

        print(f"\n  IMD TEMPERATURE GRID ({IMD_TEMP_RES} deg):")
        self.temp_cells = self._find_cells(watershed_gdf, IMD_TEMP_RES)
        print(f"\n    Total temperature cells: {len(self.temp_cells)}")

    def download_rainfall(self) -> None:
        line("[AREA-WEIGHTED] DOWNLOADING IMD RAINFALL DATA")
        for idx, cell in enumerate(self.rainfall_cells):
            print(f"\n  Cell {idx + 1}/{len(self.rainfall_cells)}: ({cell.lat:.2f}, {cell.lon:.2f}) weight={cell.weight:.3f}")
            centre_lat = cell.lat + cell.resolution / 2.0
            centre_lon = cell.lon + cell.resolution / 2.0
            chunks = []
            for year in range(self.start_year, self.end_year + 1):
                print(f"    Year {year}: requesting IMD rain dataset ... ", end="", flush=True)
                try:
                    rain_ds = self._load_imd_xarray("rain", year)
                    print("OK")
                    pt = rain_ds.sel(lat=centre_lat, lon=centre_lon, method="nearest")
                    df_year = pt["rain"].to_dataframe(name="precip_mm").reset_index()
                    df_year = ensure_datetime_column(df_year, "time")
                    df_year = df_year[df_year["time"].dt.year == year]
                    df_year["precip_mm"] = df_year["precip_mm"].replace(-999.0, 0.0).clip(lower=0.0)
                    chunks.append(df_year[["time", "precip_mm"]])
                    print(f"      extracted {len(df_year)} days @ ({centre_lat:.3f},{centre_lon:.3f})")
                except Exception as exc:
                    print("FAILED")
                    fail(f"{year}: {str(exc)[:120]}")
            if chunks:
                cell.data = pd.concat(chunks, ignore_index=True).sort_values("time")
                annual_avg = cell.data["precip_mm"].sum() / max(1, self.end_year - self.start_year + 1)
                ok(f"{len(cell.data)} days; annual mean = {annual_avg:.1f} mm")
            else:
                cell.data = pd.DataFrame(columns=["time", "precip_mm"])

    def download_temperature(self) -> None:
        line("[AREA-WEIGHTED] DOWNLOADING IMD TEMPERATURE DATA")
        if not self.temp_cells:
            print("  No temperature cells found. Skipping.")
            return
        for idx, cell in enumerate(self.temp_cells):
            print(f"\n  Cell {idx + 1}/{len(self.temp_cells)}: ({cell.lat:.2f}, {cell.lon:.2f}) weight={cell.weight:.3f}")
            centre_lat = cell.lat + cell.resolution / 2.0
            centre_lon = cell.lon + cell.resolution / 2.0
            chunks = []
            for year in range(self.start_year, self.end_year + 1):
                print(f"    Year {year}: requesting IMD tmax/tmin ... ", end="", flush=True)
                try:
                    tmax_ds = self._load_imd_xarray("tmax", year)
                    tmin_ds = self._load_imd_xarray("tmin", year)
                    print("OK")

                    df_tmax = (
                        tmax_ds.sel(lat=centre_lat, lon=centre_lon, method="nearest")["tmax"]
                        .to_dataframe(name="tmax_c")
                        .reset_index()
                    )
                    df_tmin = (
                        tmin_ds.sel(lat=centre_lat, lon=centre_lon, method="nearest")["tmin"]
                        .to_dataframe(name="tmin_c")
                        .reset_index()
                    )
                    df_tmax = ensure_datetime_column(df_tmax, "time")
                    df_tmin = ensure_datetime_column(df_tmin, "time")
                    df_tmax = df_tmax[df_tmax["time"].dt.year == year]
                    df_tmin = df_tmin[df_tmin["time"].dt.year == year]
                    df_tmax["tmax_c"] = df_tmax["tmax_c"].replace(-999.0, np.nan)
                    df_tmin["tmin_c"] = df_tmin["tmin_c"].replace(-999.0, np.nan)
                    chunks.append(df_tmax[["time", "tmax_c"]].merge(df_tmin[["time", "tmin_c"]], on="time", how="left"))
                    print(f"      extracted {len(df_tmax)} days @ ({centre_lat:.3f},{centre_lon:.3f})")
                except Exception as exc:
                    print("FAILED")
                    fail(f"{year}: {str(exc)[:120]}")
            if chunks:
                df = pd.concat(chunks, ignore_index=True).sort_values("time")
                df["tmax_c"] = df["tmax_c"].interpolate().ffill().bfill()
                df["tmin_c"] = df["tmin_c"].interpolate().ffill().bfill()
                cell.data = df
                ok(f"{len(df)} days")
            else:
                cell.data = pd.DataFrame(columns=["time", "tmax_c", "tmin_c"])
            print(f"  Finished temperature cell {idx + 1}/{len(self.temp_cells)}")

    def calculate_area_weighted_weather(self) -> Optional[pd.DataFrame]:
        line("[AREA-WEIGHTED] CALCULATING WATERSHED AVERAGE")
        date_values = set()
        for cell in self.rainfall_cells:
            if cell.data is not None and not cell.data.empty:
                date_values.update(pd.to_datetime(cell.data["time"]).tolist())
        if not date_values:
            fail("No rainfall data available.")
            return None

        df = pd.DataFrame({"time": sorted(date_values)})
        df["precip_mm"] = 0.0
        df["tmax_c"] = 0.0
        df["tmin_c"] = 0.0

        precip_weight = 0.0
        print("\n  Calculating area-weighted rainfall...")
        for cell in self.rainfall_cells:
            if cell.data is not None and not cell.data.empty:
                merged = df[["time"]].merge(cell.data[["time", "precip_mm"]], on="time", how="left")
                df["precip_mm"] += merged["precip_mm"].fillna(0.0) * cell.weight
                precip_weight += cell.weight
        if 0 < precip_weight < 0.999:
            df["precip_mm"] /= precip_weight

        temp_weight = 0.0
        print("  Calculating area-weighted temperature...")
        for cell in self.temp_cells:
            if cell.data is not None and not cell.data.empty:
                merged = df[["time"]].merge(cell.data[["time", "tmax_c", "tmin_c"]], on="time", how="left")
                df["tmax_c"] += merged["tmax_c"].fillna(0.0) * cell.weight
                df["tmin_c"] += merged["tmin_c"].fillna(0.0) * cell.weight
                temp_weight += cell.weight
        if temp_weight > 0:
            if temp_weight < 0.999:
                df["tmax_c"] /= temp_weight
                df["tmin_c"] /= temp_weight
        else:
            df["tmax_c"] = np.nan
            df["tmin_c"] = np.nan

        df["tavg_c"] = (df["tmax_c"] + df["tmin_c"]) / 2.0
        n_years = max(1, self.end_year - self.start_year + 1)
        ok(f"{len(df)} days computed")
        print(f"    Annual precip: {df['precip_mm'].sum() / n_years:.1f} mm")
        print(f"    Mean Tmax:     {df['tmax_c'].mean():.1f} deg C")
        print(f"    Mean Tmin:     {df['tmin_c'].mean():.1f} deg C")
        return df


# ============================================================================ #
# NASA POWER DOWNLOADER
# ============================================================================ #

class NASAPOWERDownloader:
    """Download solar radiation, wind speed and relative humidity."""

    BASE_URL = "https://power.larc.nasa.gov/api/temporal/daily/point"

    def __init__(self, start_year: int, end_year: int):
        self.start_year = start_year
        self.end_year = end_year

    def download_for_station(self, lat: float, lon: float) -> Optional[pd.DataFrame]:
        params = {
            "parameters": "ALLSKY_SFC_SW_DWN,WS2M,RH2M",
            "community": "AG",
            "longitude": lon,
            "latitude": lat,
            "start": f"{self.start_year}0101",
            "end": f"{self.end_year}1231",
            "format": "JSON",
        }
        try:
            resp = requests.get(self.BASE_URL, params=params, timeout=60)
            resp.raise_for_status()
            raw = resp.json()["properties"]["parameter"]
            df = pd.DataFrame(raw)
            df.index = pd.to_datetime(df.index)
            df = df.reset_index()
            df.columns = ["time", "solar_mj_m2", "wind_ms", "humidity_pct"]
            df = df.replace(-999.0, np.nan).interpolate().ffill().bfill()
            return df
        except Exception as exc:
            fail(f"NASA POWER error: {exc}")
            return None


# ============================================================================ #
# MAIN SWAT DOWNLOADER
# ============================================================================ #

class SWATDataDownloaderReal:
    """Top-level orchestrator for SWAT weather input data."""

    def __init__(
        self,
        shp_path: str,
        dem_path: str,
        output_dir: str,
        start_year: int,
        end_year: int,
        imd_raw_folder: str = IMD_RAW_FOLDER,
        output_file_prefix: str = "Thoubal_SWAT",
        area_name: str = "Lower Thoubal River Catchment",
        station_name_prefix: str = "Station",
        num_bands: int = 5,
        stations_per_band: int = 1,
    ):
        self.shp_path = shp_path
        self.dem_path = dem_path
        self.output_dir = output_dir
        self.start_year = start_year
        self.end_year = end_year
        self.imd_raw_folder = imd_raw_folder
        self.output_file_prefix = safe_filename_stem(output_file_prefix, "SWAT_Weather")
        self.area_name = area_name.strip() or "Watershed"
        self.station_name_prefix = safe_filename_stem(station_name_prefix, "Station")
        self.num_bands = num_bands
        self.stations_per_band = stations_per_band

        self.dirs = {
            "swat": os.path.join(output_dir, "SWAT_Weather_Files"),
            "excel": os.path.join(output_dir, "Excel_Data"),
            "reports": os.path.join(output_dir, "Reports"),
        }
        for folder in self.dirs.values():
            os.makedirs(folder, exist_ok=True)

        line("SWAT WEATHER DATA DOWNLOADER")
        self.watershed_gdf = gpd.read_file(shp_path)
        require_crs(self.watershed_gdf, "Watershed shapefile")
        self.watershed_wgs84 = self.watershed_gdf.to_crs("EPSG:4326")
        self.bounds = self.watershed_wgs84.total_bounds

        self.dem_proc = DEMProcessor(dem_path)
        self.dem_proc.load_dem()
        self.elev_stats = self.dem_proc.extract_watershed_stats(self.watershed_gdf)
        self.elev_bands = self.dem_proc.create_elevation_bands(self.watershed_gdf, num_bands)

        self.watershed_area = self._calc_area()
        self.mean_elev = self.elev_stats["mean"] if self.elev_stats else MANUAL_ELEVATION

        print(f"\n  Watershed area:  {self.watershed_area:.2f} km2")
        print(f"  Mean elevation:  {self.mean_elev:.0f} m")
        if self.elev_stats:
            print(f"  Elevation range: {self.elev_stats['range']:.0f} m")

        self.imd_downloader = AreaWeightedIMDDownloader(imd_raw_folder, start_year, end_year)
        self.nasa_downloader = NASAPOWERDownloader(start_year, end_year)

        self.stations: list[dict] = []
        self.all_data: dict[int, pd.DataFrame] = {}
        self.base_weather: Optional[pd.DataFrame] = None

    def _calc_area(self) -> float:
        return float(self.watershed_gdf.to_crs("EPSG:6933").area.sum() / 1_000_000.0)

    def generate_stations(self) -> None:
        line("[STEP 1] GENERATING ELEVATION-BASED STATIONS")
        bounds = self.watershed_wgs84.total_bounds
        ws_poly = geometry_union(self.watershed_wgs84)
        rng = np.random.default_rng(seed=42)

        for _, band in self.elev_bands.iterrows():
            print(f"\n  Band {int(band['Band'])}: {band['Elev_Min_m']:.0f}-{band['Elev_Max_m']:.0f} m")
            for _ in range(self.stations_per_band):
                placed = False
                for _attempt in range(200):
                    lon = float(rng.uniform(bounds[0], bounds[2]))
                    lat = float(rng.uniform(bounds[1], bounds[3]))
                    if not ws_poly.contains(Point(lon, lat)):
                        continue
                    elev = self.dem_proc.extract_point_elevation(lat, lon) or float(band["Elev_Center_m"])
                    if float(band["Elev_Min_m"]) <= elev <= float(band["Elev_Max_m"]):
                        sid = len(self.stations) + 1
                        station = {
                            "ID": sid,
                            "Name": f"{self.station_name_prefix}_{sid}",
                            "LAT": lat,
                            "LON": lon,
                            "ELEV": float(elev),
                            "BAND": int(band["Band"]),
                            "TEMP_ADJ": float(band["Temp_Adjustment_C"]),
                            "PRECIP_MULT": float(band["Precip_Multiplier"]),
                        }
                        self.stations.append(station)
                        print(f"    Station {sid}: {lat:.4f}, {lon:.4f}, {elev:.0f} m")
                        placed = True
                        break
                if not placed:
                    warn(f"Could not place station in band {int(band['Band'])}; skipping.")
        ok(f"{len(self.stations)} stations generated")

    def download_watershed_weather(self) -> bool:
        line("[STEP 2] DOWNLOADING AREA-WEIGHTED IMD WEATHER")
        self.imd_downloader.identify_overlapping_cells(self.watershed_wgs84)
        self.imd_downloader.download_rainfall()
        self.imd_downloader.download_temperature()
        self.base_weather = self.imd_downloader.calculate_area_weighted_weather()
        if self.base_weather is None:
            fail("Weather download failed.")
            return False
        return True

    def generate_station_weather(self) -> None:
        line("[STEP 3] GENERATING STATION-SPECIFIC WEATHER")
        if self.base_weather is None:
            raise RuntimeError("Base weather has not been downloaded.")
        self._create_location_files()

        for station in self.stations:
            sid = station["ID"]
            df = self.base_weather.copy()
            elev_diff = float(station["ELEV"]) - float(self.mean_elev)
            t_adj = TEMP_LAPSE_RATE * elev_diff / 1000.0
            p_mult = 1.0 + (PRECIP_LAPSE_RATE / 100.0) * max(0.0, elev_diff) / 100.0

            df["tmax_c"] = df["tmax_c"].fillna(25.0) + t_adj
            df["tmin_c"] = df["tmin_c"].fillna(15.0) + t_adj
            df["tavg_c"] = df["tavg_c"].fillna(20.0) + t_adj
            df["precip_mm"] = df["precip_mm"].fillna(0.0) * p_mult

            nasa = self.nasa_downloader.download_for_station(station["LAT"], station["LON"])
            if nasa is not None:
                df = df.merge(nasa, on="time", how="left")
            else:
                df["solar_mj_m2"] = 15.0
                df["wind_ms"] = 2.0
                df["humidity_pct"] = 60.0

            for column, default in (("solar_mj_m2", 15.0), ("wind_ms", 2.0), ("humidity_pct", 60.0)):
                df[column] = df[column].interpolate().ffill().bfill().fillna(default)

            self.all_data[sid] = df
            self._save_standard_swat_files(df, sid)
            print(f"    Station {sid}: {station['ELEV']:.0f} m  T adj: {t_adj:+.1f} C  P mult: {p_mult:.2f}x")
        ok("All station weather files created")

    def _create_location_files(self) -> None:
        for prefix in ("pcp", "tmp", "slr", "wnd", "hmd"):
            path = os.path.join(self.dirs["swat"], f"{prefix}.txt")
            with open(path, "w", encoding="utf-8") as fh:
                fh.write("ID,NAME,LAT,LONG,ELEVATION\n")
                for station in self.stations:
                    swat_station_name = f"{prefix}{station['ID']}"
                    fh.write(
                        f"{station['ID']},{swat_station_name},"
                        f"{station['LAT']:.4f},{station['LON']:.4f},{station['ELEV']:.1f}\n"
                    )
        ok("Location files created")

    def _save_standard_swat_files(self, df: pd.DataFrame, sid: int) -> None:
        swat = self.dirs["swat"]

        def write_series(name: str, column: str, default: float) -> None:
            with open(os.path.join(swat, f"{name}{sid}.txt"), "w", encoding="utf-8") as fh:
                for value in df[column].fillna(default):
                    fh.write(f"{value:.2f}\n")

        write_series("pcp", "precip_mm", 0.0)
        write_series("slr", "solar_mj_m2", 15.0)
        write_series("wnd", "wind_ms", 2.0)
        write_series("hmd", "humidity_pct", 60.0)

        tmp_path = os.path.join(swat, f"tmp{sid}.txt")
        with open(tmp_path, "w", encoding="utf-8") as fh:
            for _, row in df.iterrows():
                tmax = row["tmax_c"] if pd.notna(row["tmax_c"]) else 25.0
                tmin = row["tmin_c"] if pd.notna(row["tmin_c"]) else 15.0
                fh.write(f"{tmax:.2f},{tmin:.2f}\n")

    def _weather_with_power_columns(self) -> Optional[pd.DataFrame]:
        if self.base_weather is None:
            return None
        df = self.base_weather.copy()
        if self.all_data:
            first_sid = next(iter(self.all_data))
            station_df = self.all_data[first_sid]
            for col_name in ("solar_mj_m2", "wind_ms", "humidity_pct"):
                if col_name in station_df.columns and col_name not in df.columns:
                    df = df.merge(station_df[["time", col_name]], on="time", how="left")
        return df

    def create_excel(self) -> None:
        line("[STEP 4] CREATING MASTER EXCEL WORKBOOK")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            warn("openpyxl not found. Falling back to basic xlsxwriter export.")
            self._create_excel_basic()
            return

        out_path = os.path.join(self.dirs["excel"], f"{self.output_file_prefix}_Complete.xlsx")
        wb = Workbook()
        wb.remove(wb.active)

        navy = "1F3864"
        steel = "2E75B6"
        gold = "F4B942"
        grey = "D6DCE4"
        row_odd = "EBF3FB"
        row_even = "FFFFFF"
        green = "E2EFDA"
        red = "FCE4D6"
        blue = "DDEBF7"
        yellow = "FFFF99"
        orange = "FCE4D6"
        purple = "EAD1DC"

        def fill(color: str) -> PatternFill:
            return PatternFill("solid", fgColor=color)

        def thin_border() -> Border:
            side = Side(style="thin", color="AAAAAA")
            return Border(left=side, right=side, top=side, bottom=side)

        def center() -> Alignment:
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

        def left() -> Alignment:
            return Alignment(horizontal="left", vertical="center", wrap_text=True)

        def title(ws, text: str, cells: str, size: int = 14) -> None:
            ws.merge_cells(cells)
            c = ws[cells.split(":")[0]]
            c.value = text
            c.font = Font(bold=True, color="FFFFFF", size=size, name="Calibri")
            c.fill = fill(navy)
            c.alignment = center()

        def write_header(ws, row: int, headers: list[str], colors: Optional[list[str]] = None) -> None:
            for col_idx, header in enumerate(headers, 1):
                c = ws.cell(row=row, column=col_idx, value=header)
                c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
                c.fill = fill(colors[col_idx - 1] if colors else navy)
                c.alignment = center()
                c.border = thin_border()
            ws.row_dimensions[row].height = 32

        def write_row(ws, row: int, values: list, colors: Optional[list[str]] = None, formats: Optional[list[str]] = None) -> None:
            band = row_odd if row % 2 else row_even
            for col_idx, value in enumerate(values, 1):
                c = ws.cell(row=row, column=col_idx, value=value)
                c.fill = fill(colors[col_idx - 1] if colors else band)
                c.font = Font(color="1A1A2E", size=9, name="Calibri")
                c.alignment = left()
                c.border = thin_border()
                if formats and formats[col_idx - 1]:
                    c.number_format = formats[col_idx - 1]

        # Cover
        ws = wb.create_sheet("Cover")
        ws.sheet_view.showGridLines = False
        title(ws, "SWAT HYDROLOGICAL MODEL - WEATHER DATA PACKAGE", "A1:H3", size=18)
        ws.merge_cells("A4:H4")
        ws["A4"] = self.area_name
        ws["A4"].font = Font(bold=True, color=gold, size=13, name="Calibri")
        ws["A4"].fill = fill(navy)
        ws["A4"].alignment = center()

        cover_rows = [
            ("Watershed Shapefile", os.path.basename(self.shp_path), ""),
            ("Total Area", f"{self.watershed_area:.2f} km2", ""),
            ("Mean Elevation", f"{self.mean_elev:.0f} m a.s.l.", ""),
            ("Elevation Range", f"{self.elev_stats['range']:.0f} m" if self.elev_stats else "N/A", ""),
            ("Simulation Period", f"{self.start_year}-{self.end_year}", ""),
            ("Elevation Bands", self.num_bands, ""),
            ("Stations Generated", len(self.stations), ""),
            ("Temp Lapse Rate", f"{TEMP_LAPSE_RATE} deg C / 1000 m", "Environmental lapse rate"),
            ("Precip Lapse Rate", f"{PRECIP_LAPSE_RATE} % / 100 m", "Orographic enhancement"),
            ("Created", datetime.now().strftime("%Y-%m-%d %H:%M"), ""),
        ]
        write_header(ws, 6, ["Field", "Value", "Note"], [steel, steel, steel])
        for idx, row_data in enumerate(cover_rows, 7):
            write_row(ws, idx, list(row_data), [grey, row_odd if idx % 2 else row_even, row_odd if idx % 2 else row_even])
        for col, width in zip("ABCDEFGH", [24, 24, 36, 24, 24, 24, 18, 18]):
            ws.column_dimensions[col].width = width

        # Data sources
        ws2 = wb.create_sheet("Data_Sources")
        ws2.sheet_view.showGridLines = False
        title(ws2, "COMPLETE DATA SOURCE CITATIONS - Thoubal SWAT Project", "A1:H2")
        headers = ["Variable", "Product Name", "Provider", "Resolution", "Period", "URL/API", "SWAT File(s)", "Notes"]
        write_header(ws2, 4, headers)
        source_rows = [
            ("Rainfall", "IMD Gridded Daily Rainfall", "IMD Pune", "0.25 deg daily", "1901-present", "imdpune.gov.in / imdlib", "pcp*.txt", "Area-weighted over watershed."),
            ("Tmax", "IMD Gridded Tmax", "IMD Pune", "1.0 deg daily", "1951-present", "imdpune.gov.in / imdlib", "tmp*.txt", "Area-weighted and lapse-rate adjusted."),
            ("Tmin", "IMD Gridded Tmin", "IMD Pune", "1.0 deg daily", "1951-present", "imdpune.gov.in / imdlib", "tmp*.txt", "Area-weighted and lapse-rate adjusted."),
            ("Solar", "NASA POWER ALLSKY_SFC_SW_DWN", "NASA LaRC", "Point daily", "1981-present", "power.larc.nasa.gov", "slr*.txt", "MJ/m2/day."),
            ("Wind", "NASA POWER WS2M", "NASA LaRC", "Point daily", "1981-present", "power.larc.nasa.gov", "wnd*.txt", "m/s at 2 m."),
            ("Humidity", "NASA POWER RH2M", "NASA LaRC", "Point daily", "1981-present", "power.larc.nasa.gov", "hmd*.txt", "Relative humidity percent."),
        ]
        for idx, row_data in enumerate(source_rows, 5):
            write_row(ws2, idx, list(row_data))
            ws2.row_dimensions[idx].height = 44
        for col, width in zip("ABCDEFGH", [18, 34, 24, 18, 16, 42, 24, 44]):
            ws2.column_dimensions[col].width = width
        ws2.freeze_panes = "A5"
        ws2.auto_filter.ref = f"A4:H{4 + len(source_rows)}"

        # Elevation bands
        ws3 = wb.create_sheet("Elevation_Bands")
        ws3.sheet_view.showGridLines = False
        title(ws3, "ELEVATION BANDS - DEM-derived", "A1:G2")
        headers = ["Band", "Elev Min (m)", "Elev Max (m)", "Centre (m)", "Area Fraction", "Temp Adj (C)", "Precip Mult"]
        write_header(ws3, 3, headers, [steel] * 7)
        formats = ["0", "0", "0", "0", "0.0%", "+0.00;-0.00;0.00", "0.00"]
        for idx, (_, row) in enumerate(self.elev_bands.iterrows(), 4):
            values = [
                int(row["Band"]),
                row["Elev_Min_m"],
                row["Elev_Max_m"],
                row["Elev_Center_m"],
                row["Area_Fraction"],
                row["Temp_Adjustment_C"],
                row["Precip_Multiplier"],
            ]
            write_row(ws3, idx, values, formats=formats)
        for col, width in zip("ABCDEFG", [8, 15, 15, 15, 15, 15, 15]):
            ws3.column_dimensions[col].width = width

        # Weather stations
        ws4 = wb.create_sheet("Weather_Stations")
        ws4.sheet_view.showGridLines = False
        title(ws4, "WEATHER STATIONS - Elevation-band Representative Points", "A1:H2")
        headers = ["Station ID", "Name", "Latitude", "Longitude", "Elevation (m)", "Band", "Temp Adj (C)", "Precip Mult"]
        write_header(ws4, 3, headers, [steel] * 8)
        for idx, station in enumerate(self.stations, 4):
            values = [
                station["ID"],
                station["Name"],
                station["LAT"],
                station["LON"],
                station["ELEV"],
                station["BAND"],
                station["TEMP_ADJ"],
                station["PRECIP_MULT"],
            ]
            write_row(ws4, idx, values, formats=["0", "@", "0.0000", "0.0000", "0", "0", "+0.0;-0.0;0.0", "0.00"])
        for col, width in zip("ABCDEFGH", [12, 16, 14, 14, 14, 8, 14, 14]):
            ws4.column_dimensions[col].width = width

        # Watershed average
        bw = self._weather_with_power_columns()
        if bw is not None:
            ws5 = wb.create_sheet("Watershed_Avg")
            ws5.sheet_view.showGridLines = False
            title(ws5, "AREA-WEIGHTED WATERSHED DAILY WEATHER - IMD + NASA POWER", "A1:H2")
            headers = [
                "Date",
                "Rainfall\n(mm)",
                "Tmax\n(C)",
                "Tmin\n(C)",
                "Tavg\n(C)",
                "Solar Rad.\n(MJ/m2/day)",
                "Wind\n(m/s)",
                "Rel. Humidity\n(%)",
            ]
            colors = [navy, green, red, blue, blue, yellow, orange, purple]
            write_header(ws5, 3, headers, colors)
            formats = ["DD-MMM-YYYY", "0.00", "0.0", "0.0", "0.0", "0.00", "0.00", "0.0"]
            for idx, (_, row) in enumerate(bw.iterrows(), 4):
                date_val = row["time"].to_pydatetime() if hasattr(row["time"], "to_pydatetime") else row["time"]
                values = [
                    date_val,
                    round(row.get("precip_mm", 0.0), 2),
                    round(row.get("tmax_c", np.nan), 1),
                    round(row.get("tmin_c", np.nan), 1),
                    round(row.get("tavg_c", np.nan), 1),
                    round(row.get("solar_mj_m2", np.nan), 2) if pd.notna(row.get("solar_mj_m2", np.nan)) else "",
                    round(row.get("wind_ms", np.nan), 2) if pd.notna(row.get("wind_ms", np.nan)) else "",
                    round(row.get("humidity_pct", np.nan), 1) if pd.notna(row.get("humidity_pct", np.nan)) else "",
                ]
                write_row(ws5, idx, values, colors, formats)
            ws5.freeze_panes = "A4"
            ws5.auto_filter.ref = f"A3:H{3 + len(bw)}"
            for col, width in zip("ABCDEFGH", [14, 12, 10, 10, 10, 16, 10, 14]):
                ws5.column_dimensions[col].width = width

        # Per-station sheets
        for sid, df in self.all_data.items():
            station = next(s for s in self.stations if s["ID"] == sid)
            ws = wb.create_sheet(f"Station_{sid}")
            ws.sheet_view.showGridLines = False
            title(
                ws,
                f"Station {sid} - {station['Name']} | Lat {station['LAT']:.4f}, Lon {station['LON']:.4f}, Elev {station['ELEV']:.0f} m",
                "A1:H2",
                size=11,
            )
            headers = [
                "Date",
                "Rainfall\n(mm/day)",
                "Tmax\n(C)",
                "Tmin\n(C)",
                "Tavg\n(C)",
                "Solar Rad.\n(MJ/m2/day)",
                "Wind Speed\n(m/s)",
                "Rel. Humidity\n(%)",
            ]
            colors = [navy, green, red, blue, blue, yellow, orange, purple]
            write_header(ws, 3, headers, colors)
            formats = ["DD-MMM-YYYY", "0.00", "0.0", "0.0", "0.0", "0.00", "0.00", "0.0"]
            for idx, (_, row) in enumerate(df.iterrows(), 4):
                date_val = row["time"].to_pydatetime() if hasattr(row["time"], "to_pydatetime") else row["time"]
                values = [
                    date_val,
                    round(row.get("precip_mm", 0.0), 2),
                    round(row.get("tmax_c", np.nan), 1),
                    round(row.get("tmin_c", np.nan), 1),
                    round(row.get("tavg_c", np.nan), 1),
                    round(row.get("solar_mj_m2", 15.0), 2),
                    round(row.get("wind_ms", 2.0), 2),
                    round(row.get("humidity_pct", 60.0), 1),
                ]
                write_row(ws, idx, values, colors, formats)
            ws.freeze_panes = "A4"
            ws.auto_filter.ref = f"A3:H{3 + len(df)}"
            for col, width in zip("ABCDEFGH", [14, 14, 10, 10, 10, 16, 12, 14]):
                ws.column_dimensions[col].width = width

        # Monthly stats
        if bw is not None and not bw.empty:
            ws7 = wb.create_sheet("Monthly_Stats")
            ws7.sheet_view.showGridLines = False
            title(ws7, "MONTHLY CLIMATOLOGY - Basin Mean", "A1:H2")
            headers = ["Month", "Total Rainfall\n(mm)", "Mean Tmax\n(C)", "Mean Tmin\n(C)", "Mean Tavg\n(C)", "Mean Solar\n(MJ/m2/day)", "Mean Wind\n(m/s)", "Mean RH\n(%)"]
            colors = [navy, green, red, blue, blue, yellow, orange, purple]
            write_header(ws7, 3, headers, colors)
            monthly = bw.copy()
            monthly["month"] = monthly["time"].dt.month
            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            n_years = max(1, self.end_year - self.start_year + 1)
            for month_num in range(1, 13):
                grp = monthly[monthly["month"] == month_num]
                values = [
                    months[month_num - 1],
                    round(grp["precip_mm"].sum() / n_years, 1) if not grp.empty else "",
                    round(grp["tmax_c"].mean(), 1) if not grp.empty else "",
                    round(grp["tmin_c"].mean(), 1) if not grp.empty else "",
                    round(grp["tavg_c"].mean(), 1) if not grp.empty else "",
                    round(grp["solar_mj_m2"].mean(), 2) if "solar_mj_m2" in grp and not grp.empty else "",
                    round(grp["wind_ms"].mean(), 2) if "wind_ms" in grp and not grp.empty else "",
                    round(grp["humidity_pct"].mean(), 1) if "humidity_pct" in grp and not grp.empty else "",
                ]
                write_row(ws7, month_num + 3, values, colors, ["@", "0.0", "0.0", "0.0", "0.0", "0.00", "0.00", "0.0"])
            for col, width in zip("ABCDEFGH", [10, 18, 14, 14, 14, 18, 14, 14]):
                ws7.column_dimensions[col].width = width

        # Annual summary
        if self.all_data:
            ws8 = wb.create_sheet("Annual_Summary")
            ws8.sheet_view.showGridLines = False
            title(ws8, "ANNUAL SUMMARY - All Stations", "A1:I2")
            headers = ["Year", "Station", "Ann. Precip\n(mm)", "Mean Tmax\n(C)", "Mean Tmin\n(C)", "Mean Tavg\n(C)", "Mean Solar\n(MJ/m2/day)", "Mean Wind\n(m/s)", "Mean RH\n(%)"]
            colors = [navy, navy, green, red, blue, blue, yellow, orange, purple]
            write_header(ws8, 3, headers, colors)
            row_idx = 4
            for sid, df_station in self.all_data.items():
                station = next(s for s in self.stations if s["ID"] == sid)
                df_year = df_station.copy()
                df_year["year"] = df_year["time"].dt.year
                for year, grp in df_year.groupby("year"):
                    values = [
                        int(year),
                        station["Name"],
                        round(grp["precip_mm"].sum(), 1),
                        round(grp["tmax_c"].mean(), 1),
                        round(grp["tmin_c"].mean(), 1),
                        round(grp["tavg_c"].mean(), 1),
                        round(grp["solar_mj_m2"].mean(), 2),
                        round(grp["wind_ms"].mean(), 2),
                        round(grp["humidity_pct"].mean(), 1),
                    ]
                    write_row(ws8, row_idx, values, colors, ["0", "@", "0.0", "0.0", "0.0", "0.0", "0.00", "0.00", "0.0"])
                    row_idx += 1
            ws8.freeze_panes = "A4"
            ws8.auto_filter.ref = f"A3:I{row_idx - 1}"
            for col, width in zip("ABCDEFGHI", [8, 14, 16, 14, 14, 14, 18, 14, 14]):
                ws8.column_dimensions[col].width = width

        tab_colours = {
            "Cover": navy,
            "Data_Sources": steel,
            "Elevation_Bands": "70AD47",
            "Weather_Stations": "ED7D31",
            "Watershed_Avg": "4472C4",
            "Monthly_Stats": "5B9BD5",
            "Annual_Summary": "A9D18E",
        }
        for sheet_name in wb.sheetnames:
            if sheet_name in tab_colours:
                wb[sheet_name].sheet_properties.tabColor = tab_colours[sheet_name]
            elif sheet_name.startswith("Station_"):
                wb[sheet_name].sheet_properties.tabColor = "ED7D31"

        wb.save(out_path)
        ok(f"Excel workbook: {out_path}")
        ok(f"Sheets: {', '.join(wb.sheetnames)}")

    def _create_excel_basic(self) -> None:
        out = os.path.join(self.dirs["excel"], f"{self.output_file_prefix}_Weather.xlsx")
        with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
            if self.base_weather is not None:
                self.base_weather.to_excel(writer, sheet_name="Watershed_Avg", index=False)
            for sid, df in self.all_data.items():
                df.to_excel(writer, sheet_name=f"Station_{sid}"[:31], index=False)
        ok(f"Basic Excel: {out}")

    def create_report(self) -> None:
        line("[STEP 5] CREATING FINAL REPORT")
        path = os.path.join(self.dirs["reports"], f"{self.output_file_prefix}_Weather_Report.txt")
        lines = [
            "=" * 80,
            "SWAT WEATHER DATA REPORT",
            "=" * 80,
            "",
            f"Watershed : {self.area_name}",
            f"Area      : {self.watershed_area:.2f} km2",
            f"Mean elev : {self.mean_elev:.0f} m",
            f"Period    : {self.start_year}-{self.end_year}",
            "",
            "WEATHER INPUTS",
            "  Rainfall      : IMD 0.25 deg gridded, area-weighted",
            "  Temperature   : IMD 1.0 deg gridded, area-weighted",
            "  Solar/Wind/RH : NASA POWER",
            "",
            "OUTPUTS",
            "  SWAT weather files : pcp/tmp/slr/wnd/hmd station files",
            "  Excel workbook     : formatted daily and summary weather tables",
            "=" * 80,
        ]
        with open(path, "w", encoding="utf-8") as fh:
            fh.write("\n".join(lines) + "\n")
        ok(f"Report: {path}")

    def run_all(self) -> None:
        t0 = time.time()
        self.generate_stations()
        if self.download_watershed_weather():
            self.generate_station_weather()
            self.create_excel()
        self.create_report()
        elapsed = time.time() - t0
        line("WEATHER DATA DOWNLOAD COMPLETE")
        print(f"  Time: {elapsed / 60.0:.1f} min")
        print(f"  Output: {self.output_dir}")


# ============================================================================ #
# ENTRY POINT
# ============================================================================ #

def main() -> int:
    line("SWAT WEATHER DATA DOWNLOADER")
    print("Lower Thoubal River Catchment, Manipur")

    if not HAS_IMDLIB:
        fail("imdlib is not installed. Run: pip install imdlib")
        return 1

    if not os.path.exists(WATERSHED_SHP):
        print(f"\nERROR: Shapefile not found:\n  {WATERSHED_SHP}")
        return 1

    print("\nConfiguration:")
    print(f"  Watershed       : {WATERSHED_SHP}")
    print(f"  DEM             : {DEM_PATH}")
    print(f"  Output          : {OUTPUT_DIR}")
    print(f"  Period          : {START_YEAR}-{END_YEAR}")
    print(f"  Elevation bands : {NUM_ELEVATION_BANDS}")
    print("\nWeather data sources:")
    print("  Rainfall        : IMD 0.25 deg gridded")
    print("  Temperature     : IMD 1.0 deg gridded")
    print("  Solar/Wind/RH   : NASA POWER")

    proceed = input("\nProceed? (Y/N): ").strip().upper()
    if proceed != "Y":
        print("Cancelled.")
        return 0

    downloader = SWATDataDownloaderReal(
        shp_path=WATERSHED_SHP,
        dem_path=DEM_PATH,
        output_dir=OUTPUT_DIR,
        start_year=START_YEAR,
        end_year=END_YEAR,
        imd_raw_folder=IMD_RAW_FOLDER,
        output_file_prefix="Thoubal_SWAT",
        area_name="Lower Thoubal River Catchment",
        station_name_prefix="Station",
        num_bands=NUM_ELEVATION_BANDS,
        stations_per_band=STATIONS_PER_BAND,
    )

    downloader.run_all()

    line("WEATHER DATA DOWNLOADED SUCCESSFULLY")
    print(f"  Weather files : {downloader.dirs['swat']}")
    print(f"  Excel workbook: {downloader.dirs['excel']}")
    print(f"  Report        : {downloader.dirs['reports']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
